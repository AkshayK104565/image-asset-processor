"""
core/engine.py  –  Pattern Image Asset Processor  (Pillow-only, no ImageMagick)

WHY NO IMAGEMAGICK:
  Render free tier has 512 MB RAM. ImageMagick spawns a subprocess per image
  and can spike to 150-300 MB per call. With parallel workers this causes OOM.

  Pure Pillow stays in-process, processes a 2000x2000 JPG in ~15 MB of RAM,
  and is 3-5x faster than spawning a subprocess.

WHAT PILLOW HANDLES:
  - JPEG, PNG, WebP, BMP, GIF, TIFF → square white-padded JPG
  - EXIF auto-rotation
  - Alpha removal (transparent → white background)
  - sRGB colour space
  - PDF passthrough (copied as-is; image→PDF via ghostscript if available,
    otherwise saved as JPG with .pdf extension warning)
"""

import io, os, re, shutil, tempfile, threading, time, urllib.parse, zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import requests
from PIL import Image, ImageOps, ExifTags

# ── constants ─────────────────────────────────────────────
DEFAULT_WORKERS = 8          # conservative default for free tier
HTTP_TIMEOUT    = 30
MAX_RETRIES     = 2

SHEET_IMAGE = "Image Downloader"
SHEET_PXM   = "PXM advanced search"
SHEET_INST  = "Instructions"

IMAGE_EXTS = {".jpg",".jpeg",".png",".bmp",".gif",".tif",".tiff",".webp",".jfif"}
CT_EXT = {
    "application/pdf":".pdf","image/jpeg":".jpg","image/png":".png",
    "image/webp":".webp","image/gif":".gif","image/bmp":".bmp",
    "image/tiff":".tif","image/jfif":".jpg",
}

# Excel highlight styles
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")
RED_FONT    = Font(color="CC0000", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")

# ── thread-safe helpers ───────────────────────────────────
_cnt = 0; _cnt_lock = threading.Lock()
_sess = None; _sess_lock = threading.Lock()

def _uid():
    global _cnt
    with _cnt_lock:
        _cnt += 1
        return f"img_{int(time.time()*1000)}_{_cnt}"

def _get_sess():
    global _sess
    if _sess is None:
        with _sess_lock:
            if _sess is None:
                s = requests.Session()
                s.headers.update({"User-Agent": "Mozilla/5.0"})
                _sess = s
    return _sess

def _trim(v): return str(v).strip() if v is not None else ""
def _clean(s):
    for c in r'\/:*?"<>|': s = s.replace(c, "")
    return s or "UNKNOWN"
def _ext(p): return os.path.splitext(p)[1].lower()
def _normext(e):
    e = e.strip().lower()
    return ("." + e) if e and not e.startswith(".") else e
def _rm(p):
    try:
        if p and os.path.exists(p): os.remove(p)
    except: pass

def parse_dim(txt):
    txt = txt.strip().lower().replace(" ", "")
    if "x" in txt:
        parts = txt.split("x")
        if len(parts) != 2: return 0
        try:
            w, h = int(parts[0]), int(parts[1])
            return max(w, h) if w > 0 and h > 0 else 0
        except: return 0
    try:
        v = int(txt); return v if v > 0 else 0
    except: return 0

def is_valid_master_id(s):
    s = s.strip()
    return len(s) == 8 and all(c.isalnum() for c in s)

def find_magick():
    """No longer used for images. Kept only as a health-check indicator."""
    return "pillow"   # always available


# ── PILLOW IMAGE PROCESSING ───────────────────────────────

def _fix_orientation(img: Image.Image) -> Image.Image:
    """Apply EXIF rotation so images aren't sideways."""
    try:
        exif = img._getexif()
        if exif is None:
            return img
        orientation_key = next(
            (k for k, v in ExifTags.TAGS.items() if v == "Orientation"), None
        )
        if orientation_key is None:
            return img
        orientation = exif.get(orientation_key)
        rotations = {3: 180, 6: 270, 8: 90}
        if orientation in rotations:
            img = img.rotate(rotations[orientation], expand=True)
    except Exception:
        pass
    return img


def _to_rgb_white(img: Image.Image) -> Image.Image:
    """Convert any mode to RGB, compositing transparency onto white."""
    if img.mode in ("RGBA", "LA", "PA"):
        background = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "PA":
            img = img.convert("RGBA")
        if img.mode == "LA":
            img = img.convert("RGBA")
        background.paste(img, mask=img.split()[-1])
        return background
    if img.mode == "P":
        img = img.convert("RGBA")
        background = Image.new("RGB", img.size, (255, 255, 255))
        background.paste(img, mask=img.split()[-1])
        return background
    if img.mode != "RGB":
        img = img.convert("RGB")
    return img


def _to_square_jpg_pillow(src_path: str, dst_path: str,
                           min_size: int, max_size: int):
    """
    Pure-Pillow conversion to white-padded square JPG.
    Returns (ok: bool, action: str)
    action ∈ {'UPSIZED', 'DOWNSIZED', 'ORIGINAL', ''}
    Memory usage: ~15-40 MB per image (vs 150-300 MB for ImageMagick subprocess)
    """
    try:
        with Image.open(src_path) as raw:
            img = _fix_orientation(raw.copy())

        img = _to_rgb_white(img)
        w, h = img.size

        if w < min_size or h < min_size:
            target = min_size
            action = "UPSIZED"
            # Resize up so the shorter side reaches target
            ratio = target / min(w, h)
            new_w, new_h = int(w * ratio), int(h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)
        elif w > max_size or h > max_size:
            target = max_size
            action = "DOWNSIZED"
            # Resize down so the longer side becomes target
            ratio = target / max(w, h)
            new_w, new_h = int(w * ratio), int(h * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)
        else:
            target = max(w, h)
            action = "ORIGINAL"

        # Pad to square with white
        w2, h2 = img.size
        square_size = max(w2, h2)
        if w2 != h2:
            canvas = Image.new("RGB", (square_size, square_size), (255, 255, 255))
            paste_x = (square_size - w2) // 2
            paste_y = (square_size - h2) // 2
            canvas.paste(img, (paste_x, paste_y))
            img = canvas

        img.save(dst_path, format="JPEG", quality=92,
                 optimize=True, progressive=True)
        img.close()
        return True, action

    except Exception as e:
        return False, ""


def _image_to_jpg_pillow(src_path: str, dst_path: str) -> bool:
    """Convert any image to plain JPG (used for SDS pipeline)."""
    try:
        with Image.open(src_path) as raw:
            img = _fix_orientation(raw.copy())
        img = _to_rgb_white(img)
        img.save(dst_path, format="JPEG", quality=92)
        img.close()
        return True
    except Exception:
        return False


def _jpg_to_pdf(jpg_path: str, pdf_path: str) -> bool:
    """
    Convert JPG to single-page PDF.
    Uses Pillow's built-in PDF writer — no ghostscript needed.
    """
    try:
        with Image.open(jpg_path) as img:
            rgb = img.convert("RGB")
            rgb.save(pdf_path, format="PDF", resolution=150)
        return os.path.isfile(pdf_path)
    except Exception:
        return False


# ── URL / extension helpers ───────────────────────────────
def _gdrive(url):
    url = url.strip(); fid = ""
    if "drive.usercontent.google.com" in url:
        fid = _qp(url, "id")
    else:
        m = re.search(r"/file/d/([^/]+)", url)
        if m: fid = m.group(1)
        elif any(x in url for x in ("drive.google.com/open",
                                     "drive.google.com/uc",
                                     "docs.google.com/uc")):
            fid = _qp(url, "id")
    if fid:
        return f"https://drive.usercontent.google.com/u/0/uc?id={fid}&export=download"
    return url

def _qp(url, p):
    return (urllib.parse.parse_qs(urllib.parse.urlparse(url).query).get(p) or [""])[0]

def _det_ext(orig, ct, src):
    e = _ext(orig)
    if e: return _normext(e)
    ct = ct.lower().split(";")[0].strip()
    for k, v in CT_EXT.items():
        if k in ct: return v
    u = src.split("?")[0].split("#")[0]
    e = _ext(u)
    return _normext(e) if e and len(e) <= 10 else ""

def _sniff_pdf(p):
    try:
        with open(p, "rb") as f: return f.read(4) == b"%PDF"
    except: return False

def _download(url, tmp):
    final = _gdrive(url); last = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            r = _get_sess().get(final, timeout=HTTP_TIMEOUT,
                                allow_redirects=True, stream=True)
            r.raise_for_status()
            ct = r.headers.get("Content-Type", "").lower()
            if "text/html" in ct or "application/xhtml" in ct:
                raise ValueError("URL returned an HTML page, not a file")
            orig = ""
            cd = r.headers.get("Content-Disposition", "")
            if cd:
                m = re.search(r'filename="?([^";]+)"?', cd)
                if m: orig = m.group(1).strip()
            e = _det_ext(orig, ct, final) or ".bin"
            lp = os.path.join(tmp, _uid() + e)
            with open(lp, "wb") as fout:
                for chunk in r.iter_content(chunk_size=131072):
                    if chunk: fout.write(chunk)
            return lp, ct, orig
        except Exception as exc:
            last = exc
            if attempt < MAX_RETRIES: time.sleep(1.5 * (attempt + 1))
    raise last


# ── single-file worker ────────────────────────────────────
def _process_one(task):
    url    = task["url"]
    fname  = task["final_name"]
    atype  = task["asset_type"]
    out    = task["out_dir"]
    mn     = task["min_size"]
    mx     = task["max_size"]
    tmp    = task["tmp_dir"]
    row    = task["row"]
    col    = task["col"]

    res = {"final_name": fname, "ok": False, "is_non_jpg": False,
           "resize_action": "", "ext": ".jpg" if atype != "sds" else ".pdf",
           "row": row, "col": col, "url": url, "error": ""}
    tp = ""
    try:
        tp, ct, _ = _download(url, tmp)
        se = _ext(tp)

        if atype in ("image", "label"):
            if se not in (".jpg", ".jpeg"): res["is_non_jpg"] = True
            if se in IMAGE_EXTS or "image/" in ct:
                dst = os.path.join(out, fname)
                ok, action = _to_square_jpg_pillow(tp, dst, mn, mx)
                res["ok"] = ok; res["resize_action"] = action
                if not ok: res["error"] = "Pillow conversion failed (corrupt or unsupported image)"
            else:
                res["error"] = f"Not a recognised image format ({se})"

        elif atype == "sds":
            dst = os.path.join(out, fname)
            if se == ".pdf" or "application/pdf" in ct or _sniff_pdf(tp):
                shutil.copy2(tp, dst)
                res["ok"] = os.path.isfile(dst)
                res["ext"] = ".pdf"
            elif se in IMAGE_EXTS or "image/" in ct:
                tmp_jpg = os.path.join(tmp, _uid() + ".jpg")
                if _image_to_jpg_pillow(tp, tmp_jpg):
                    res["ok"] = _jpg_to_pdf(tmp_jpg, dst)
                    _rm(tmp_jpg)
                res["ext"] = ".pdf"
                if not res["ok"]: res["error"] = "SDS image→PDF conversion failed"
            else:
                res["error"] = f"SDS is not a PDF or image ({se})"

    except Exception as e:
        res["error"] = str(e)
    finally:
        _rm(tp)
    return res


# ── Excel helpers ─────────────────────────────────────────
def _fc(ws, name):
    for cell in ws[1]:
        if _trim(cell.value) == name: return cell.column
    return 0

def _fcp(ws, prefix):
    up = prefix.upper()
    for cell in ws[1]:
        if _trim(cell.value).upper().startswith(up): return cell.column
    return 0

def _lcp(ws, start, prefix):
    up = prefix.upper(); end = start
    for c in range(start + 1, ws.max_column + 1):
        if _trim(ws.cell(1, c).value).upper().startswith(up): end = c
        else: break
    return end

def _mlr(ws, cols):
    mx = 0
    for c in cols:
        if c and c > 0:
            for r in range(ws.max_row, 1, -1):
                if ws.cell(r, c).value not in (None, ""):
                    if r > mx: mx = r
                    break
    return max(mx, 1)

def _rha(ws, row, cis, cie, cls_, cle, csds):
    for s, e in [(cis, cie), (cls_, cle)]:
        if s and e and e >= s:
            for c in range(s, e + 1):
                if _trim(ws.cell(row, c).value): return True
    return bool(csds and _trim(ws.cell(row, csds).value))


# ── validation ────────────────────────────────────────────
def validate_workbook(wb, upload_choice):
    errors = []
    needed = [SHEET_IMAGE, SHEET_INST]
    if upload_choice == "1": needed.append(SHEET_PXM)
    for s in needed:
        if s not in wb.sheetnames: errors.append(f"Sheet '{s}' not found.")
    if errors: return errors
    ws = wb[SHEET_IMAGE]; wi = wb[SHEET_INST]
    cm   = _fc(ws, "Master ID");  cmpn = _fc(ws, "MPN")
    cc   = _fc(ws, "Country Code"); ca  = _fc(ws, "ASIN")
    cis  = _fcp(ws, "Image URL"); cie  = _lcp(ws, cis,  "Image URL")  if cis  else 0
    cls_ = _fcp(ws, "Label Image URL"); cle = _lcp(ws, cls_, "Label Image URL") if cls_ else 0
    csds = _fc(ws, "SDS URL")
    cols = [c for c in [cm, cmpn, cc, ca, cis, cie, cls_, cle, csds] if c]
    lr   = _mlr(ws, cols)
    if lr < 2: errors.append("No data rows in 'Image Downloader'."); return errors
    valid_cc = set()
    for r in range(2, wi.max_row + 1):
        code = _trim(wi.cell(r, 2).value)
        if code: valid_cc.add(code.upper())
    bad = []
    for r in range(2, lr + 1):
        if not _rha(ws, r, cis, cie, cls_, cle, csds): continue
        if upload_choice == "1":
            for col, name in [(cm,"Master ID"),(cmpn,"MPN"),(cc,"Country Code")]:
                if not col or not _trim(ws.cell(r, col).value):
                    bad.append(f"Row {r}: missing {name}")
            mid = _trim(ws.cell(r, cm).value) if cm else ""
            if mid and not is_valid_master_id(mid):
                bad.append(f"Row {r}: Master ID '{mid}' must be 8 alphanumeric chars")
        else:
            for col, name in [(cc,"Country Code"),(ca,"ASIN")]:
                if not col or not _trim(ws.cell(r, col).value):
                    bad.append(f"Row {r}: missing {name}")
        code = _trim(ws.cell(r, cc).value) if cc else ""
        if code and valid_cc and code.upper() not in valid_cc:
            bad.append(f"Row {r}: Country Code '{code}' not in Instructions sheet")
    errors.extend(bad[:20])
    if len(bad) > 20: errors.append(f"…and {len(bad)-20} more errors.")
    return errors


# ── highlighted Excel builder ─────────────────────────────
def _build_highlighted_excel(excel_path, failed_cells):
    """
    failed_cells: list of (row, col, url, error_reason)
    Returns bytes of workbook with red cells on failed URLs +
    yellow row highlight + "Failure Reason" column appended.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception:
        return None
    if SHEET_IMAGE not in wb.sheetnames:
        return None
    ws = wb[SHEET_IMAGE]

    failed_rows   = set(r for r, c, u, e in failed_cells)
    failed_coords = {(r, c): e for r, c, u, e in failed_cells}

    for row in range(2, ws.max_row + 1):
        if row not in failed_rows: continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row, col_idx)
            if (row, col_idx) in failed_coords:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            else:
                ft = cell.fill.fill_type
                if ft is None or ft == "none":
                    cell.fill = YELLOW_FILL

    reason_col = ws.max_column + 1
    hdr = ws.cell(1, reason_col)
    hdr.value = "⚠ Failure Reason"
    hdr.fill  = PatternFill("solid", fgColor="C62828")
    hdr.font  = Font(bold=True, color="FFFFFF", name="Calibri")
    ws.column_dimensions[get_column_letter(reason_col)].width = 60

    row_reasons = defaultdict(list)
    for r, c, u, e in failed_cells:
        col_letter = get_column_letter(c)
        short_url  = (u[:55] + "…") if len(u) > 55 else u
        row_reasons[r].append(f"[Col {col_letter}] {e} — {short_url}")

    for r, reasons in row_reasons.items():
        cell = ws.cell(r, reason_col)
        cell.value = " | ".join(reasons)
        cell.fill  = RED_FILL
        cell.font  = RED_FONT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── main entry point ──────────────────────────────────────
def run_job(excel_path, upload_choice, min_size, max_size,
            num_workers=DEFAULT_WORKERS, progress_cb=None):
    """
    Pure-Pillow image processing. No ImageMagick, no subprocesses.
    num_workers: 1-32 parallel download+convert threads.
    """
    num_workers = max(1, min(int(num_workers), 32))

    wb   = openpyxl.load_workbook(excel_path, data_only=True)
    ws   = wb[SHEET_IMAGE]; wi = wb[SHEET_INST]
    cm   = _fc(ws,"Master ID");   cmpn = _fc(ws,"MPN")
    cc   = _fc(ws,"Country Code"); ca  = _fc(ws,"ASIN")
    cis  = _fcp(ws,"Image URL");   cie  = _lcp(ws,cis,"Image URL")   if cis  else 0
    cls_ = _fcp(ws,"Label Image URL"); cle = _lcp(ws,cls_,"Label Image URL") if cls_ else 0
    csds = _fc(ws,"SDS URL")
    cols = [c for c in [cm,cmpn,cc,ca,cis,cie,cls_,cle,csds] if c]
    lr   = _mlr(ws, cols)

    jid     = _uid()
    tmp_dir = tempfile.mkdtemp(prefix=f"tmp_{jid}_")
    out_dir = tempfile.mkdtemp(prefix=f"out_{jid}_")

    tasks = []; seen = set(); dupes = 0

    for row in range(2, lr + 1):
        if not _rha(ws, row, cis, cie, cls_, cle, csds): continue
        mid     = _trim(ws.cell(row,cm).value)   if cm   else ""
        mpn     = _clean(_trim(ws.cell(row,cmpn).value) if cmpn else "")
        country = _trim(ws.cell(row,cc).value)   if cc   else ""
        asin    = _trim(ws.cell(row,ca).value)   if ca   else ""
        row_out = out_dir
        if upload_choice == "2" and country:
            row_out = os.path.join(out_dir, country)
            os.makedirs(row_out, exist_ok=True)
        base = {"min_size":min_size,"max_size":max_size,
                "tmp_dir":tmp_dir,"out_dir":row_out,"row":row}

        def _add(url, fname, atype, col):
            nonlocal dupes
            if not url: return
            if fname.upper() in seen: dupes += 1; return
            seen.add(fname.upper())
            tasks.append({**base,"url":url,"final_name":fname,
                          "asset_type":atype,"col":col})

        if cis and cie:
            for c in range(cis, cie+1):
                url = _trim(ws.cell(row,c).value); idx = c - cis + 1
                fname = (f"{mid}_{mpn}_{country}_ISP{idx:02d}.jpg" if upload_choice=="1"
                         else (f"{asin}.MAIN.jpg" if idx==1 else f"{asin}.PT{idx-1:02d}.jpg"))
                _add(url, fname, "image", c)

        if cls_ and cle:
            for c in range(cls_, cle+1):
                url = _trim(ws.cell(row,c).value); idx = c - cls_ + 1
                fname = (f"{mid}_{mpn}_{country}_LBL{idx:02d}.jpg" if upload_choice=="1"
                         else f"{asin}.PS{idx:02d}.jpg")
                _add(url, fname, "label", c)

        if csds:
            url = _trim(ws.cell(row,csds).value)
            fname = (f"{mid}_{mpn}_{country}_SDS.pdf" if upload_choice=="1"
                     else f"{asin}_SDS.pdf")
            _add(url, fname, "sds", csds)

    total = len(tasks); downloaded = 0; failed = 0
    upsized = 0; downsized = 0; kept = 0
    ftypes = defaultdict(int); errs = []
    failed_cells = []

    pxm_lock = threading.Lock()
    pxm_entries = []; pxm_seen = set(); non_jpg = set()

    def rec_pxm(fname, row, is_nj):
        k = fname.upper()
        with pxm_lock:
            if is_nj: non_jpg.add(row)
            if k not in pxm_seen: pxm_seen.add(k); pxm_entries.append((row,fname))

    done = 0
    with ThreadPoolExecutor(max_workers=num_workers) as ex:
        futs = {ex.submit(_process_one, t): t for t in tasks}
        for fut in as_completed(futs):
            r = fut.result(); done += 1
            if r["ok"]:
                downloaded += 1; ftypes[r["ext"]] += 1
                a = r.get("resize_action","")
                if a=="UPSIZED":   upsized   += 1
                elif a=="DOWNSIZED": downsized += 1
                elif a=="ORIGINAL":  kept      += 1
                if upload_choice=="1": rec_pxm(r["final_name"],r["row"],r.get("is_non_jpg",False))
            else:
                failed += 1
                err_msg = r.get("error","Unknown error")
                errs.append({"row":r["row"],"col":r["col"],"file":r["final_name"],
                              "url":r.get("url",""),"error":err_msg})
                failed_cells.append((r["row"],r["col"],r.get("url",""),err_msg))
            if progress_cb: progress_cb(done, total, r["final_name"])

    # Build ZIP
    zip_path = os.path.join(tmp_dir, f"images_{jid}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root,_,files in os.walk(out_dir):
            for fn in files:
                fp = os.path.join(root,fn)
                zf.write(fp, os.path.relpath(fp, out_dir))
    shutil.rmtree(out_dir, ignore_errors=True)

    # Build highlighted Excel
    highlighted_excel_path = None
    if failed_cells:
        hbytes = _build_highlighted_excel(excel_path, failed_cells)
        if hbytes:
            highlighted_excel_path = os.path.join(tmp_dir, f"failed_links_{jid}.xlsx")
            with open(highlighted_excel_path, "wb") as f: f.write(hbytes)

    prio = [f for (rn,f) in pxm_entries if rn in non_jpg]
    norm = [f for (rn,f) in pxm_entries if rn not in non_jpg]

    return {
        "zip_path":               zip_path,
        "highlighted_excel_path": highlighted_excel_path,
        "downloaded":  downloaded, "failed":    failed,
        "duplicates":  dupes,      "total":     total,
        "upsized":     upsized,    "downsized": downsized, "kept_orig": kept,
        "min_size":    min_size,   "max_size":  max_size,
        "file_types":  dict(ftypes), "pxm_list": prio+norm,
        "errors_log":  errs,       "error":     None,
    }
